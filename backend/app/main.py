from __future__ import annotations

import json
import re
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Annotated

from fastapi import Depends, FastAPI, File, HTTPException, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from sqlalchemy import inspect, or_, text
from sqlalchemy.orm import Session

from .db import Base, SessionLocal, engine, get_db
from .excel import workbook_rows
from .models import AuditLog, ControlSettings, LicenseItem, User
from .schemas import (
    AuditLogRead,
    ControlSettingsRead,
    ControlSettingsUpdate,
    CustomFieldDefinition,
    DashboardResponse,
    HealthResponse,
    ImportResult,
    LicenseCreate,
    LicenseRead,
    LicenseUpdate,
    LoginRequest,
    PasswordChange,
    TokenResponse,
    UserCreate,
    UserRead,
    UserRoleUpdate,
)
from .email_service import notify_owner_alert, notify_owner_update, notify_status_change, notify_new_email_subscription
from .security import create_access_token, get_current_user, hash_password, require_roles, verify_password
from .services import apply_payload_to_item, dashboard_payload, enrich_item, filter_items, item_to_risk_payload, log_audit
from .settings import settings


app = FastAPI(title=settings.app_name, version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[origin.strip() for origin in settings.cors_origins.split(",") if origin.strip()],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def bootstrap_database() -> None:
    Base.metadata.create_all(bind=engine)


DEFAULT_CONTROL_LISTS = {
    "category_options": [
        "SSL Certificate",
        "Server Management",
        "Endpoint Security",
        "Virtualization",
        "Operating System",
        "Network Appliance",
        "Backup",
        "Monitoring",
        "SaaS",
        "Stack-X",
        "Tickting Solution",
        "Other",
    ],
    "item_type_options": [
        "License",
        "Subscription",
        "Certificate",
        "Support Contract",
        "Warranty",
        "EOL/Lifecycle",
        "Maintenance",
        "Domian Subscription",
        "Other",
    ],
    "environment_options": ["Production", "DR", "Test", "UAT", "Office", "Cloud", "Branch", "Shared"],
    "renewal_cycle_options": ["Monthly", "Quarterly", "Semi-Annual", "Annual", "Multi-Year", "One-Time", "N/A"],
    "auto_renew_options": ["Yes", "No"],
    "priority_options": ["Low", "Medium", "High", "Critical"],
    "currency_options": ["USD", "EUR", "EGP", "SAR", "AED", "GBP", "Other"],
}


def serialize_options(options: list[str]) -> str:
    return "\n".join(item.strip() for item in options if item.strip())


def parse_options(text_value: str | None, fallback: list[str]) -> list[str]:
    if not text_value:
        return fallback[:]
    values = [value.strip() for value in text_value.splitlines() if value.strip()]
    return values or fallback[:]


def sanitize_option_list(values: list[str], fallback: list[str]) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()
    for value in values:
        item = str(value).strip()
        if not item:
            continue
        key = item.lower()
        if key in seen:
            continue
        seen.add(key)
        normalized.append(item)
    return normalized or fallback[:]


def parse_custom_rules(text_value: str | None) -> list[dict[str, object]]:
    if not text_value:
        return []
    try:
        parsed = json.loads(text_value)
    except json.JSONDecodeError:
        return []
    return parsed if isinstance(parsed, list) else []


def serialize_custom_rules(rules: list[object]) -> str:
    return json.dumps(rules, ensure_ascii=True)


def parse_custom_field_definitions(text_value: str | None) -> list[dict[str, object]]:
    if not text_value:
        return []
    try:
        parsed = json.loads(text_value)
    except json.JSONDecodeError:
        return []
    return parsed if isinstance(parsed, list) else []


def serialize_custom_field_definitions(defs: list[object]) -> str:
    return json.dumps(defs, ensure_ascii=True)


def ensure_control_settings_columns() -> None:
    inspector = inspect(engine)
    if "control_settings" not in inspector.get_table_names():
        return

    existing_columns = {column["name"] for column in inspector.get_columns("control_settings")}
    required_columns = {
        "category_options": serialize_options(DEFAULT_CONTROL_LISTS["category_options"]),
        "item_type_options": serialize_options(DEFAULT_CONTROL_LISTS["item_type_options"]),
        "environment_options": serialize_options(DEFAULT_CONTROL_LISTS["environment_options"]),
        "renewal_cycle_options": serialize_options(DEFAULT_CONTROL_LISTS["renewal_cycle_options"]),
        "auto_renew_options": serialize_options(DEFAULT_CONTROL_LISTS["auto_renew_options"]),
        "priority_options": serialize_options(DEFAULT_CONTROL_LISTS["priority_options"]),
        "currency_options": serialize_options(DEFAULT_CONTROL_LISTS["currency_options"]),
        "custom_rules": "[]",
        "custom_field_definitions": "[]",
    }

    with engine.begin() as conn:
        for column_name, default_value in required_columns.items():
            if column_name in existing_columns:
                continue
            escaped_default = default_value.replace("'", "''")
            conn.execute(
                text(
                    f"ALTER TABLE control_settings ADD COLUMN {column_name} TEXT NOT NULL DEFAULT '{escaped_default}'"
                )
            )


def ensure_license_items_columns() -> None:
    inspector = inspect(engine)
    if "license_items" not in inspector.get_table_names():
        return
    existing_columns = {col["name"] for col in inspector.get_columns("license_items")}
    with engine.begin() as conn:
        if "custom_fields" not in existing_columns:
            conn.execute(text("ALTER TABLE license_items ADD COLUMN custom_fields TEXT NOT NULL DEFAULT '{}'"))
        if "email" not in existing_columns:
            conn.execute(text("ALTER TABLE license_items ADD COLUMN email VARCHAR(255) NOT NULL DEFAULT ''"))


def seed_users(db: Session) -> None:
    seed_data = [
        (settings.demo_admin_email, settings.demo_admin_password, "admin", "Admin User"),
        (settings.demo_ops_email, settings.demo_ops_password, "ops", "Operations User"),
        (settings.demo_viewer_email, settings.demo_viewer_password, "viewer", "Read Only User"),
    ]
    for email, password, role, full_name in seed_data:
        existing = db.query(User).filter(User.email == email).one_or_none()
        if existing is None:
            db.add(User(email=email, password_hash=hash_password(password), role=role, full_name=full_name))
    db.commit()


def seed_control_settings(db: Session) -> None:
    existing = db.query(ControlSettings).order_by(ControlSettings.id.asc()).first()
    if existing is None:
        db.add(
            ControlSettings(
                urgent_days_threshold=settings.default_alert_days_urgent,
                review_days_threshold=settings.default_alert_days_review,
                eol_soon_threshold=90,
                default_reminder_lead_time=settings.default_alert_days_review,
                base_currency="USD",
                template_version="1.0",
                category_options=serialize_options(DEFAULT_CONTROL_LISTS["category_options"]),
                item_type_options=serialize_options(DEFAULT_CONTROL_LISTS["item_type_options"]),
                environment_options=serialize_options(DEFAULT_CONTROL_LISTS["environment_options"]),
                renewal_cycle_options=serialize_options(DEFAULT_CONTROL_LISTS["renewal_cycle_options"]),
                auto_renew_options=serialize_options(DEFAULT_CONTROL_LISTS["auto_renew_options"]),
                priority_options=serialize_options(DEFAULT_CONTROL_LISTS["priority_options"]),
                currency_options=serialize_options(DEFAULT_CONTROL_LISTS["currency_options"]),
                custom_rules="[]",
                custom_field_definitions="[]",
            )
        )
        db.commit()


@app.on_event("startup")
def startup_event() -> None:
    bootstrap_database()
    ensure_control_settings_columns()
    ensure_license_items_columns()
    db = SessionLocal()
    try:
        seed_users(db)
        seed_control_settings(db)
    finally:
        db.close()


def get_control_settings_record(db: Session) -> ControlSettings:
    control = db.query(ControlSettings).order_by(ControlSettings.id.asc()).first()
    if control is None:
        control = ControlSettings(
            urgent_days_threshold=settings.default_alert_days_urgent,
            review_days_threshold=settings.default_alert_days_review,
            eol_soon_threshold=90,
            default_reminder_lead_time=settings.default_alert_days_review,
            base_currency="USD",
            template_version="1.0",
            category_options=serialize_options(DEFAULT_CONTROL_LISTS["category_options"]),
            item_type_options=serialize_options(DEFAULT_CONTROL_LISTS["item_type_options"]),
            environment_options=serialize_options(DEFAULT_CONTROL_LISTS["environment_options"]),
            renewal_cycle_options=serialize_options(DEFAULT_CONTROL_LISTS["renewal_cycle_options"]),
            auto_renew_options=serialize_options(DEFAULT_CONTROL_LISTS["auto_renew_options"]),
            priority_options=serialize_options(DEFAULT_CONTROL_LISTS["priority_options"]),
            currency_options=serialize_options(DEFAULT_CONTROL_LISTS["currency_options"]),
            custom_rules="[]",
            custom_field_definitions="[]",
        )
        db.add(control)
        db.commit()
        db.refresh(control)
    return control


def control_settings_to_read(control: ControlSettings) -> ControlSettingsRead:
    return ControlSettingsRead(
        urgent_days_threshold=control.urgent_days_threshold,
        review_days_threshold=control.review_days_threshold,
        eol_soon_threshold=control.eol_soon_threshold,
        default_reminder_lead_time=control.default_reminder_lead_time,
        base_currency=control.base_currency,
        template_version=control.template_version,
        category_options=parse_options(control.category_options, DEFAULT_CONTROL_LISTS["category_options"]),
        item_type_options=parse_options(control.item_type_options, DEFAULT_CONTROL_LISTS["item_type_options"]),
        environment_options=parse_options(control.environment_options, DEFAULT_CONTROL_LISTS["environment_options"]),
        renewal_cycle_options=parse_options(control.renewal_cycle_options, DEFAULT_CONTROL_LISTS["renewal_cycle_options"]),
        auto_renew_options=parse_options(control.auto_renew_options, DEFAULT_CONTROL_LISTS["auto_renew_options"]),
        priority_options=parse_options(control.priority_options, DEFAULT_CONTROL_LISTS["priority_options"]),
        currency_options=parse_options(control.currency_options, DEFAULT_CONTROL_LISTS["currency_options"]),
        custom_rules=parse_custom_rules(control.custom_rules),
        custom_field_definitions=parse_custom_field_definitions(control.custom_field_definitions),
    )


def get_thresholds(control: ControlSettings) -> dict[str, object]:
    return {
        "urgent_days_threshold": control.urgent_days_threshold,
        "review_days_threshold": control.review_days_threshold,
        "eol_soon_threshold": control.eol_soon_threshold,
        "default_reminder_lead_time": control.default_reminder_lead_time,
        "custom_rules": parse_custom_rules(control.custom_rules),
    }


def serialize_item(item: LicenseItem, thresholds: dict[str, object]) -> LicenseRead:
    return LicenseRead.model_validate(enrich_item(item, thresholds=thresholds))


def get_items(db: Session, query: str | None = None, status_filter: str | None = None, category: str | None = None) -> list[LicenseItem]:
    items = db.query(LicenseItem).all()
    return filter_items(items, query, status_filter, category)

@app.get("/")
def root():
    frontend_dist_path = Path(__file__).resolve().parent.parent.parent / "frontend" / "dist"
    if not frontend_dist_path.exists():
        frontend_dist_path = Path(__file__).resolve().parent.parent / "static"
    
    index_file = frontend_dist_path / "index.html"
    if index_file.exists():
        return FileResponse(index_file)
    return {"status": "ok"}

@app.get("/api/control-settings", response_model=ControlSettingsRead)
def control_settings_get(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ControlSettingsRead:
    _ = current_user
    control = get_control_settings_record(db)
    return control_settings_to_read(control)


@app.put("/api/control-settings", response_model=ControlSettingsRead)
def control_settings_update(
    payload: ControlSettingsUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("admin", "ops")),
) -> ControlSettingsRead:
    _ = current_user
    control = get_control_settings_record(db)
    control.urgent_days_threshold = payload.urgent_days_threshold
    control.review_days_threshold = payload.review_days_threshold
    control.eol_soon_threshold = payload.eol_soon_threshold
    control.default_reminder_lead_time = payload.default_reminder_lead_time
    control.base_currency = payload.base_currency
    control.template_version = payload.template_version
    control.category_options = serialize_options(sanitize_option_list(payload.category_options, DEFAULT_CONTROL_LISTS["category_options"]))
    control.item_type_options = serialize_options(sanitize_option_list(payload.item_type_options, DEFAULT_CONTROL_LISTS["item_type_options"]))
    control.environment_options = serialize_options(sanitize_option_list(payload.environment_options, DEFAULT_CONTROL_LISTS["environment_options"]))
    control.renewal_cycle_options = serialize_options(sanitize_option_list(payload.renewal_cycle_options, DEFAULT_CONTROL_LISTS["renewal_cycle_options"]))
    control.auto_renew_options = serialize_options(sanitize_option_list(payload.auto_renew_options, DEFAULT_CONTROL_LISTS["auto_renew_options"]))
    control.priority_options = serialize_options(sanitize_option_list(payload.priority_options, DEFAULT_CONTROL_LISTS["priority_options"]))
    control.currency_options = serialize_options(sanitize_option_list(payload.currency_options, DEFAULT_CONTROL_LISTS["currency_options"]))
    control.custom_rules = serialize_custom_rules([rule.model_dump() for rule in payload.custom_rules])
    control.custom_field_definitions = serialize_custom_field_definitions([d.model_dump() for d in payload.custom_field_definitions])
    db.commit()
    db.refresh(control)
    return control_settings_to_read(control)


@app.get("/api/health", response_model=HealthResponse)
def health() -> HealthResponse:
    return HealthResponse(status="ok")


@app.get("/api/auth/keycloak-config")
def keycloak_config() -> dict[str, Any]:
    return {
        "enabled": settings.keycloak_enabled,
        "url": settings.keycloak_url,
        "realm": settings.keycloak_realm,
        "client_id": settings.keycloak_client_id,
    }


@app.post("/api/auth/login", response_model=TokenResponse)
def login(payload: LoginRequest, db: Session = Depends(get_db)) -> TokenResponse:
    user = db.query(User).filter(User.email == payload.email).one_or_none()
    if user is None or not verify_password(payload.password, user.password_hash):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid credentials")
    token = create_access_token(subject=user.email, role=user.role)
    return TokenResponse(access_token=token, role=user.role, email=user.email)


@app.get("/api/auth/me", response_model=UserRead)
def me(current_user: User = Depends(get_current_user)) -> UserRead:
    return UserRead.model_validate(current_user)


@app.patch("/api/auth/change-password")
def change_password(
    payload: PasswordChange,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not verify_password(payload.current_password, current_user.password_hash):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Current password is incorrect")
    current_user.password_hash = hash_password(payload.new_password)
    db.commit()
    return {"detail": "Password changed successfully"}


@app.get("/api/users", response_model=list[UserRead])
def list_users(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("admin")),
) -> list[UserRead]:
    _ = current_user
    users = db.query(User).order_by(User.id.asc()).all()
    return [UserRead.model_validate(u) for u in users]


@app.patch("/api/users/{user_id}/role", response_model=UserRead)
def update_user_role(
    user_id: int,
    payload: UserRoleUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("admin")),
) -> UserRead:
    target_user = db.query(User).filter(User.id == user_id).one_or_none()
    if target_user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    
    target_user.role = payload.role
    db.commit()
    db.refresh(target_user)
    return UserRead.model_validate(target_user)


@app.post("/api/users", response_model=UserRead, status_code=status.HTTP_201_CREATED)
def create_user(
    payload: UserCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("admin")),
) -> UserRead:
    existing = db.query(User).filter(User.email == payload.email).one_or_none()
    if existing is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="A user with that username/email already exists")
    new_user = User(
        email=payload.email,
        password_hash=hash_password(payload.password),
        role=payload.role,
        full_name=payload.full_name or payload.email,
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return UserRead.model_validate(new_user)


@app.get("/api/licenses", response_model=list[LicenseRead])
def list_licenses(
    query: str | None = None,
    status_filter: str | None = None,
    category: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[LicenseRead]:
    _ = current_user
    control = get_control_settings_record(db)
    thresholds = get_thresholds(control)
    items = get_items(db, query, status_filter, category)
    return [serialize_item(item, thresholds) for item in items]


@app.post("/api/licenses", response_model=LicenseRead)
def create_license(
    payload: LicenseCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("admin", "ops")),
) -> LicenseRead:
    control = get_control_settings_record(db)
    thresholds = get_thresholds(control)
    item = LicenseItem()
    apply_payload_to_item(item, payload, thresholds=thresholds)
    db.add(item)
    db.flush()
    log_audit(db, item.id, current_user.email, "create", "*", "", item.product_service)
    db.commit()
    db.refresh(item)
    # Notify owner if a rule flagged notify-owner, or status is already Urgent/Expired on creation
    if "notify-owner" in (item.risk_flags or []):
        notify_owner_alert(item)
    elif item.status in {"Expired", "Urgent"}:
        notify_status_change(item, "")
    return serialize_item(item, thresholds)


@app.get("/api/licenses/{item_id}", response_model=LicenseRead)
def get_license(item_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)) -> LicenseRead:
    _ = current_user
    control = get_control_settings_record(db)
    thresholds = get_thresholds(control)
    item = db.query(LicenseItem).filter(LicenseItem.id == item_id).one_or_none()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    return serialize_item(item, thresholds)


@app.patch("/api/licenses/{item_id}", response_model=LicenseRead)
def update_license(
    item_id: int,
    payload: LicenseUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("admin", "ops")),
) -> LicenseRead:
    control = get_control_settings_record(db)
    thresholds = get_thresholds(control)
    item = db.query(LicenseItem).filter(LicenseItem.id == item_id).one_or_none()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")

    # Capture old email and status before applying changes
    old_email = item.email or ""
    old_status = item.status or ""
    # Capture current state before updates for audit
    before = item_to_risk_payload(item).model_dump()
    # Apply updates once
    apply_payload_to_item(item, payload, thresholds=thresholds)
    after = item_to_risk_payload(item).model_dump()
    
    # Detect changed fields as before
    changed: list[str] = []
    for field_name, old_value in before.items():
        new_value = after.get(field_name)
        if old_value != new_value:
            log_audit(db, item.id, current_user.email, "update", field_name, old_value, new_value)
            changed.append(field_name)
    
    # Determine if email was newly added
    email_added = (old_email == "") and (item.email and item.email.strip() != "")
    
    db.commit()
    db.refresh(item)
    # Send notifications after commit
    if "notify-owner" in (item.risk_flags or []):
        notify_owner_alert(item)
    elif changed:
        # If a new email was added, send subscription notice first
        if email_added:
            notify_new_email_subscription(item, current_user.email)
        # Then send status change and owner update notifications
        notify_status_change(item, old_status)
        notify_owner_update(item, current_user.email, changed)
    return serialize_item(item, thresholds)


@app.delete("/api/licenses/{item_id}")
def delete_license(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("admin")),
) -> dict[str, str]:
    item = db.query(LicenseItem).filter(LicenseItem.id == item_id).one_or_none()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    db.delete(item)
    log_audit(db, item_id, current_user.email, "delete", "*", item.product_service, "")
    db.commit()
    return {"status": "deleted"}


@app.get("/api/dashboard", response_model=DashboardResponse)
def dashboard(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)) -> DashboardResponse:
    _ = current_user
    control = get_control_settings_record(db)
    thresholds = get_thresholds(control)
    items = db.query(LicenseItem).all()
    enriched_items = [enrich_item(item, thresholds=thresholds) for item in items]
    return dashboard_payload(enriched_items, thresholds=thresholds, base_currency=control.base_currency)


@app.get("/api/alerts", response_model=list[LicenseRead])
def alerts(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)) -> list[LicenseRead]:
    _ = current_user
    control = get_control_settings_record(db)
    thresholds = get_thresholds(control)
    items = db.query(LicenseItem).filter(
        or_(LicenseItem.status.in_(["Expired", "Urgent", "Review", "Missing Expiry Info"]), LicenseItem.anomaly_score >= 15)
    ).all()
    return [serialize_item(item, thresholds) for item in filter_items(items, None, None, None)]


@app.get("/api/audit-logs", response_model=list[AuditLogRead])
def audit_logs(db: Session = Depends(get_db), current_user: User = Depends(require_roles("admin", "ops"))) -> list[AuditLogRead]:
    _ = current_user
    logs = db.query(AuditLog).order_by(AuditLog.created_at.desc()).limit(200).all()
    return [AuditLogRead.model_validate(log) for log in logs]


@app.get("/api/insights")
def insights(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)) -> dict[str, object]:
    _ = current_user
    control = get_control_settings_record(db)
    thresholds = get_thresholds(control)
    items = db.query(LicenseItem).all()
    enriched_items = [enrich_item(item, thresholds=thresholds) for item in items]
    return dashboard_payload(enriched_items, thresholds=thresholds, base_currency=control.base_currency).predictive_insights


@app.post("/api/import/xlsx", response_model=ImportResult)
def import_xlsx(
    upload: Annotated[UploadFile, File(...)],
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles("admin", "ops")),
) -> ImportResult:
    _ = current_user
    control = get_control_settings_record(db)
    thresholds = get_thresholds(control)
    upload_dir = Path(settings.upload_dir)
    upload_dir.mkdir(parents=True, exist_ok=True)
    safe_name = upload.filename or "upload.xlsx"
    # Strip any directory separators to avoid path traversal
    safe_name = Path(safe_name).name or "upload.xlsx"
    saved_path = upload_dir / safe_name
    saved_path.write_bytes(upload.file.read())

    try:
        rows, sheet_name, warnings = workbook_rows(saved_path)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=f"Could not parse workbook: {exc}") from exc

    # Build a mapping from normalised column label → custom field key for any
    # custom field definitions stored in control settings.
    custom_field_defs: list[dict] = parse_custom_field_definitions(control.custom_field_definitions)  # type: ignore[arg-type]

    # Also re-parse workbook headers to capture unmapped columns for custom fields.
    try:
        from openpyxl import load_workbook as _load_workbook
        from .excel import DEFAULT_ALIASES, choose_sheet, map_headers, normalize_key, parse_value
        _wb = _load_workbook(saved_path, data_only=True)
        _sheet = choose_sheet(_wb, DEFAULT_ALIASES)
        _all_rows = list(_sheet.iter_rows(values_only=True))
        _header_row = [str(c).strip() if c is not None else "" for c in _all_rows[0]] if _all_rows else []
        # column index → custom field key, keyed by exact label match (case-insensitive)
        custom_col_map: dict[int, str] = {}
        for _cfd in custom_field_defs:
            _label = str(_cfd.get("label", "")).strip()
            _key = str(_cfd.get("key", "")).strip()
            if not _label or not _key:
                continue
            _norm_label = normalize_key(_label)
            for _col_idx, _header in enumerate(_header_row):
                if normalize_key(_header) == _norm_label:
                    custom_col_map[_col_idx] = _key
                    break
        # Build per-row custom field values from the raw sheet rows
        # aligned with the parsed rows (skip header row)
        _data_rows = _all_rows[1:] if len(_all_rows) > 1 else []
    except Exception:
        custom_col_map = {}
        _data_rows = []
        _wb = None

    imported = updated = skipped = 0
    import_warnings: list[str] = list(warnings)

    for row_index, row in enumerate(rows):
        normalized_row = {
            key: value
            for key, value in row.items()
            if not (isinstance(value, str) and value.strip() == "")
        }

        # Inject custom field values from matched columns
        if custom_col_map and row_index < len(_data_rows):
            raw_sheet_row = _data_rows[row_index]
            existing_custom = dict(normalized_row.get("custom_fields") or {})
            for col_idx, cf_key in custom_col_map.items():
                if col_idx < len(raw_sheet_row):
                    cell_val = raw_sheet_row[col_idx]
                    if cell_val is not None and str(cell_val).strip() != "":
                        existing_custom[cf_key] = cell_val
            if existing_custom:
                normalized_row["custom_fields"] = existing_custom

        # Convert optional imported "days_to_eol" into eol_date when eol_date is not provided.
        days_to_eol_raw = normalized_row.pop("days_to_eol_source", None)
        if normalized_row.get("eol_date") in (None, "") and days_to_eol_raw not in (None, "", "-", "na", "n/a", "NA", "N/A"):
            try:
                days_value = int(float(str(days_to_eol_raw).strip()))
                normalized_row["eol_date"] = date.today() + timedelta(days=days_value)
                import_warnings.append("Row defaulted: eol_date derived from days_to_eol")
            except ValueError:
                import_warnings.append("Row warning: invalid days_to_eol value ignored")

        # Keep empty values deterministic without inventing business data.
        normalized_row.setdefault("client", "-")
        normalized_row.setdefault("category", "-")
        normalized_row.setdefault("vendor", "-")
        normalized_row.setdefault("product_service", "-")
        normalized_row.setdefault("owner", normalized_row.get("owner", "-") or "-")
        normalized_row.setdefault("renewal_owner", normalized_row.get("renewal_owner", "-") or "-")
        normalized_row.setdefault("technical_contact", normalized_row.get("technical_contact", "-") or "-")
        normalized_row.setdefault("email", normalized_row.get("email", "-") or "-")
        normalized_row.setdefault("region", normalized_row.get("region", "-") or "-")
        normalized_row.setdefault("item_type", normalized_row.get("item_type", "-") or "-")
        normalized_row.setdefault("environment", normalized_row.get("environment", "-") or "-")
        normalized_row.setdefault("renewal_cycle", normalized_row.get("renewal_cycle", "-") or "-")
        normalized_row.setdefault("notes", normalized_row.get("notes", "-") or "-")

        auto_renew_raw = normalized_row.get("auto_renew")
        if isinstance(auto_renew_raw, str):
            normalized_row["auto_renew"] = auto_renew_raw.strip().lower() in {"true", "yes", "y", "1", "auto"}

        unit_cost_raw = normalized_row.get("unit_cost")
        if isinstance(unit_cost_raw, str) and unit_cost_raw.strip().lower() in {"", "-", "na", "n/a"}:
            normalized_row["unit_cost"] = 0.0

        if "expiry_date" not in normalized_row or normalized_row.get("expiry_date") in (None, ""):
            if normalized_row.get("eol_date") not in (None, ""):
                normalized_row["expiry_date"] = normalized_row["eol_date"]
                import_warnings.append(
                    "Row defaulted: expiry_date missing, used eol_date"
                )
            elif normalized_row.get("start_date") not in (None, ""):
                normalized_row["expiry_date"] = normalized_row["start_date"]
                import_warnings.append(
                    "Row defaulted: expiry_date missing, used start_date"
                )
            else:
                normalized_row["expiry_date"] = date.today()
                notes_value = str(normalized_row.get("notes", "")).strip()
                marker = "[Missing Expiry Info]"
                normalized_row["notes"] = f"{marker} {notes_value}".strip()
                import_warnings.append(
                    "Row defaulted: expiry_date missing, used placeholder date and classified as Missing Expiry Info"
                )

        try:
            payload = LicenseCreate(**normalized_row)
        except Exception as exc:
            skipped += 1
            import_warnings.append(f"Row skipped: {exc}")
            continue

        existing = None
        if payload.license_reference:
            existing = (
                db.query(LicenseItem)
                .filter(LicenseItem.license_reference == payload.license_reference)
                .order_by(LicenseItem.id.desc())
                .first()
            )
        if existing is None:
            existing = (
                db.query(LicenseItem)
                .filter(
                    LicenseItem.client == payload.client,
                    LicenseItem.vendor == payload.vendor,
                    LicenseItem.product_service == payload.product_service,
                )
                .order_by(LicenseItem.id.desc())
                .first()
            )

        if existing is None:
            item = LicenseItem()
            apply_payload_to_item(item, payload, thresholds=thresholds)
            db.add(item)
            imported += 1
        else:
            apply_payload_to_item(existing, payload, thresholds=thresholds)
            updated += 1

    db.commit()

    # Fire notify-owner alerts for imported/updated items that have the flag
    all_items = db.query(LicenseItem).all()
    for _item in all_items:
        if "notify-owner" in (_item.risk_flags or []):
            notify_owner_alert(_item)

    return ImportResult(imported=imported, updated=updated, skipped=skipped, warnings=[f"Imported from sheet: {sheet_name}", *import_warnings])


@app.get("/api/export/xlsx")
def export_xlsx(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)) -> StreamingResponse:
    _ = current_user
    items = db.query(LicenseItem).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "License Register"
    headers = [
        "Client",
        "Region",
        "Category",
        "Item Type",
        "Vendor",
        "Product / Service",
        "Asset / Scope",
        "Environment",
        "Owner",
        "Technical Contact",
        "Email",
        "License Reference",
        "Start Date",
        "Expiry Date",
        "EOL Date",
        "Renewal Cycle",
        "Auto Renew",
        "Quantity Purchased",
        "Quantity In Use",
        "Quantity Available",
        "Utilization %",
        "Unit Cost",
        "Annual Cost",
        "Status",
        "Days to Expiry",
        "Days to EOL",
        "Priority",
        "Renewal Owner",
        "Last Reviewed",
        "Notes",
        "Source URL",
    ]
    sheet.append(headers)
    for item in items:
        sheet.append(
            [
                item.client,
                item.region,
                item.category,
                item.item_type,
                item.vendor,
                item.product_service,
                item.asset_scope,
                item.environment,
                item.owner,
                item.technical_contact,
                item.email,
                item.license_reference,
                item.start_date,
                item.expiry_date,
                item.eol_date,
                item.renewal_cycle,
                item.auto_renew,
                item.quantity_purchased,
                item.quantity_in_use,
                item.quantity_available,
                item.utilization_percent,
                item.unit_cost,
                item.annual_cost,
                item.status,
                item.days_to_expiry if item.expiry_date else None,
                item.days_to_eol if item.eol_date else None,
                item.priority,
                item.renewal_owner,
                item.last_reviewed,
                item.notes,
                item.source_url,
            ]
        )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"license-register-{datetime.now(timezone.utc).date().isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/categories")
def categories(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)) -> dict[str, list[str]]:
    _ = current_user
    items = db.query(LicenseItem.category).distinct().all()
    return {"items": sorted(category for category, in items if category)}


# Serve frontend static files if they exist (production / monolith support)
frontend_dist_path = Path(__file__).resolve().parent.parent.parent / "frontend" / "dist"
if not frontend_dist_path.exists():
    # Fallback to backend/static folder if built/packaged there
    frontend_dist_path = Path(__file__).resolve().parent.parent / "static"

if frontend_dist_path.exists():
    app.mount("/assets", StaticFiles(directory=frontend_dist_path / "assets"), name="assets")

    @app.get("/{catchall:path}")
    async def serve_frontend(catchall: str):
        if catchall.startswith("api") or catchall.startswith("docs") or catchall.startswith("openapi.json"):
            raise HTTPException(status_code=404, detail="Not Found")
        index_file = frontend_dist_path / "index.html"
        if index_file.exists():
            return FileResponse(index_file)
        raise HTTPException(status_code=404, detail="Not Found")