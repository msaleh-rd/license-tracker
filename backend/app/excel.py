from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

DATE_FIELD_NAMES = {"start_date", "expiry_date", "eol_date", "last_reviewed"}
INTEGER_FIELDS = {"quantity_purchased", "quantity_in_use", "quantity_available", "days_to_eol_source"}
FLOAT_FIELDS = {"unit_cost", "annual_cost"}

DEFAULT_ALIASES = {
    "client": ["client", "account", "customer", "tenant"],
    "region": ["region", "site", "site region", "site/region", "geo", "location"],
    "category": ["category", "type", "license_category", "class"],
    "item_type": ["item_type", "item type", "asset_type", "certificate_type", "license type"],
    "vendor": ["vendor", "provider", "supplier"],
    "product_service": ["product_service", "product / service", "product", "service", "name", "license"],
    "asset_scope": ["asset_scope", "scope", "asset", "hostname", "server", "endpoint"],
    "environment": ["environment", "env", "environment type"],
    "owner": ["primary owner", "primary_owner", "owner", "service owner", "team"],
    "technical_contact": ["technical contact", "technical_contact", "tech contact", "tech_contact", "contact"],
    "license_reference": ["license_reference", "license ref", "reference", "serial", "certificate cn"],
    "start_date": ["start_date", "commencement_date", "effective_date", "start"],
    "expiry_date": ["expiry_date", "expiration_date", "expiry", "renewal_date", "end_date"],
    "eol_date": [
        "eol_date",
        "support/eol date",
        "support eol date",
        "support end date",
        "support_end_date",
        "end_of_life_date",
        "end of life",
    ],
    "renewal_cycle": ["renewal_cycle", "renewal cycle", "cycle", "term"],
    "auto_renew": ["auto_renew", "auto renew", "auto-renew", "autorenew"],
    "quantity_purchased": [
        "quantity_purchased",
        "quantity purchased",
        "qty purchased",
        "qty_purchased",
        "purchased qty",
        "purchased",
        "quantity",
    ],
    "quantity_in_use": [
        "quantity_in_use",
        "quantity in use",
        "qty in use",
        "qty_in_use",
        "in use qty",
        "in use",
        "used",
        "consumed",
    ],
    "quantity_available": [
        "quantity_available",
        "quantity available",
        "qty available",
        "qty_available",
        "available qty",
        "available",
        "remaining",
    ],
    "unit_cost": ["unit_cost", "unit cost", "unit price", "price"],
    "annual_cost": ["annual_cost", "annual price", "annual spend", "cost"],
    "notes": ["notes", "comments", "remarks"],
    "source_url": ["source_url", "url", "source"],
    "renewal_owner": ["renewal_owner", "renewal owner", "renewal_owner_name"],
    "email": ["email", "e-mail", "notification email", "alert email", "notification_email"],
    "last_reviewed": ["last_reviewed", "last reviewed", "reviewed", "review date"],
    "days_to_eol_source": ["days to eol", "days_to_eol", "support days to eol"],
    "priority": ["priority", "criticality", "severity"],
}


def normalize_key(value: str) -> str:
    return " ".join(value.lower().replace("/", " ").replace("_", " ").split())


def parse_date_value(value: Any, workbook_epoch) -> Any:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value, workbook_epoch).date()
        except Exception:
            return value
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"true", "yes", "1"}:
        return True
    if text.lower() in {"false", "no", "0"}:
        return False
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        pass
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        return text


def parse_value(field_name: str, value: Any, workbook_epoch) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if field_name in DATE_FIELD_NAMES:
        return parse_date_value(value, workbook_epoch)
    if field_name in INTEGER_FIELDS:
        return parse_int_value(value)
    if field_name in FLOAT_FIELDS:
        return parse_float_value(value)
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if text.lower() in {"true", "yes", "1"}:
        return True
    if text.lower() in {"false", "no", "0"}:
        return False
    if text.isdigit():
        return int(text)
    try:
        return float(text)
    except ValueError:
        return text


def _clean_numeric_text(value: Any) -> str:
    text = str(value).strip()
    if not text:
        return ""
    text = text.replace(",", "").replace("_", "").replace(" ", "")
    if text.endswith("%"):
        text = text[:-1]
    return text


def parse_int_value(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value))
    cleaned = _clean_numeric_text(value)
    if cleaned.lower() in {"", "-", "na", "n/a"}:
        return 0
    try:
        return int(round(float(cleaned)))
    except ValueError:
        return 0


def parse_float_value(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = _clean_numeric_text(value)
    if cleaned.lower() in {"", "-", "na", "n/a"}:
        return 0.0
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def map_headers(headers: list[str], aliases: dict[str, list[str]]) -> dict[str, int]:
    normalized_headers = {normalize_key(str(header)): index for index, header in enumerate(headers) if str(header).strip()}
    mapping: dict[str, int] = {}
    for field_name, field_aliases in aliases.items():
        for alias in field_aliases:
            alias_key = normalize_key(alias)
            if alias_key in normalized_headers:
                mapping[field_name] = normalized_headers[alias_key]
                break
    return mapping


def score_headers(headers: list[str], aliases: dict[str, list[str]]) -> int:
    header_map = map_headers(headers, aliases)
    return len(header_map) * 10 + sum(1 for header in headers if str(header).strip())


def choose_sheet(workbook, aliases: dict[str, list[str]]):
    best_sheet = workbook.worksheets[0]
    best_score = -1
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20), values_only=True))
        if not rows:
            continue
        for row in rows[:10]:
            headers = [str(cell).strip() if cell is not None else "" for cell in row]
            score = score_headers(headers, aliases)
            if score > best_score:
                best_score = score
                best_sheet = sheet
    return best_sheet


def workbook_rows(path: Path, aliases: dict[str, list[str]] | None = None) -> tuple[list[dict[str, Any]], str, list[str]]:
    aliases = aliases or DEFAULT_ALIASES
    workbook = load_workbook(path, data_only=True)
    sheet = choose_sheet(workbook, aliases)
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], sheet.title, []

    header_row_index = 0
    best_score = -1
    for index, row in enumerate(rows[:20]):
        headers = [str(cell).strip() if cell is not None else "" for cell in row]
        score = score_headers(headers, aliases)
        non_empty = sum(1 for header in headers if header)
        score += non_empty
        if score > best_score and non_empty >= 2:
            best_score = score
            header_row_index = index

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[header_row_index]]
    header_map = map_headers(headers, aliases)
    parsed_rows: list[dict[str, Any]] = []
    warnings: list[str] = []

    for row in rows[header_row_index + 1 :]:
        row_data: dict[str, Any] = {}
        for field_name, column_index in header_map.items():
            if column_index < len(row):
                row_data[field_name] = parse_value(field_name, row[column_index], workbook.epoch)
        if any(value not in (None, "") for value in row_data.values()):
            parsed_rows.append(row_data)

    if not header_map.get("product_service"):
        warnings.append("Could not map a Product / Service column automatically.")
    if not header_map.get("expiry_date"):
        warnings.append("Could not map an Expiry Date column automatically.")

    return parsed_rows, sheet.title, warnings